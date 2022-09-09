import sys

import pcbnew

from openpyxl import Workbook, load_workbook

#0 = name
#1 = ps / cm
layer_delay = {
    "F.Cu": 58.5680,
    "In2.Cu": 70.7803,
    "B.Cu": 58.5680
}

via_delays = {
    "F.Cu_In2.Cu": 3.4740,
    "F.Cu_B.Cu": 23.9325
}

via_lengths = {
    "F.Cu_In2.Cu": 0.225,
    "F.Cu_B.Cu": 1.2
}

def mm_to_ps(layer: str, mm: float) -> float:
    return mm * (layer_delay[layer] / 10)

def get_track_flytime(tracks, netname):
    """returns the flytime of all tracks in the net
    """
    flytime = 0
    for track in tracks:
        if track.GetNet().GetNetname() == netname and type(track) == pcbnew.PCB_TRACK:
            layer = track.GetLayerName()
            length = pcbnew.ToMM(track.GetLength())
            flytime += mm_to_ps(layer, length)
    return flytime

def get_net_layers(tracks, netname):
    """returns a list of layers used by the net
    """
    layers = []
    for track in tracks:
        if track.GetNet().GetNetname() == netname and type(track) == pcbnew.PCB_TRACK:
            layer = track.GetLayerName()
            if layer not in layers:
                layers.append(layer)
    return layers

def get_via_count(tracks, netname):
    """returns the number of vias in the net
    """
    via_count = 0
    for track in tracks:
        if track.GetNet().GetNetname() == netname and type(track) == pcbnew.PCB_VIA:
            via_count += 1
    return via_count

def get_via_delay(tracks, netname):
    """returns the via delay of the net
    """
    via_delay = 0
    vias = get_via_count(tracks, netname)
    layers = get_net_layers(tracks, netname)
    for i in range(len(layers)-1):
        via_delay += via_delays[layers[i] + "_" + layers[i+1]]
    return via_delay*2

def get_track_length(tracks, netname):
    """returns the length of all tracks in the net
    """
    length = 0
    for track in tracks:
        if track.GetNet().GetNetname() == netname and type(track) == pcbnew.PCB_TRACK:
            length += pcbnew.ToMM(track.GetLength())
    return length

def get_via_length(via):



def get_via_lengths(tracks, netname):
    """returns the length of all vias in the net
    """
    length = 0.0
    bds = pcb.GetDesignSettings()
    for track in tracks:
        if track.GetNet().GetNetname() == netname and type(track) == pcbnew.PCB_VIA:
            length += get_via_length(track)
    return length

def get_sheet_col(ws, name: str) -> int:
    for col in range(1, ws.max_column+1):
        if ws.cell(row=1, column=col).value == name:
            return col - 1
    raise Exception("Could not find column \"" + name + "\"")

def update_row(ws, row, tracks):
    netname = net_prefix + row[1].value

    vias = get_via_count(tracks, netname)
    row[get_sheet_col(ws, "Vias")].value = vias

    track_flytime = get_track_flytime(tracks, netname)
    row[get_sheet_col(ws, "Track Delay")].value = round(track_flytime, 1)

    via_delay = get_via_delay(tracks, netname)
    row[get_sheet_col(ws, "Via Delay")].value = round(via_delay, 1)

    package_delay_col = get_sheet_col(ws, "Package Delay")
    package_delay = row[package_delay_col].value
    if type(package_delay) != float:
        row[package_delay_col].value = 0.0
        package_delay = 0.0

    extra_delay_col = get_sheet_col(ws, "Extra Delay")
    extra_delay = row[extra_delay_col].value
    if type(extra_delay) != float:
        row[extra_delay_col].value = 0.0
        extra_delay = 0.0

    total_delay = track_flytime + package_delay + extra_delay
    row[get_sheet_col(ws, "Total Delay")].value = round(total_delay, 1)

    layers = get_net_layers(tracks, netname)
    row[get_sheet_col(ws, "Layers")].value = ", ".join(layers)

    track_length = get_track_length(tracks, netname)
    row[get_sheet_col(ws, "Track Length")].value = round(track_length, 2)

    via_length = get_via_lengths(tracks, netname)
    row[get_sheet_col(ws, "Via Length")].value = round(via_length, 2)

    total_length = track_length + via_length
    row[get_sheet_col(ws, "Total Length")].value = round(total_length, 2)

def update_flytimes(ws, pcb, tracks):
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        if row[get_sheet_col(ws ,"Net")].value is not None:
            update_row(ws, row, tracks)

if __name__ == "__main__":

    if len(sys.argv) < 4:
        print("Usage: gen_flytimes.py <kicad_pcb> <filename> <net_prefix> <sheet_name> (optional)")
        sys.exit(1)

    pcb_path = sys.argv[1]
    wb_name = sys.argv[2]
    net_prefix = sys.argv[3]

    wb = load_workbook(wb_name)

    if len(sys.argv) >= 4:
        sheet_name = sys.argv[4]
        ws = wb[sheet_name]
    else:
        ws = wb.active

    pcb = pcbnew.LoadBoard(pcb_path)
    pcb.BuildListOfNets()
    tracks = pcb.GetTracks()
    update_flytimes(ws, pcb, tracks)
    wb.save(wb_name)
