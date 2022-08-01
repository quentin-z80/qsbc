import pcbnew

from openpyxl import Workbook, load_workbook

NET_PREFIX = "/DDR/"

#0 = name
#1 = ps / cm
layer_delay = {
    "F.Cu": 58.5680,
    "In2.Cu": 70.7803,
    "B.Cu": 58.5680
}

via_delays = {
    "F.Cu_In2.Cu": 3.4740,
    "F.Cu_B.Cu": 23.9325
}

via_lengths = {
    "F.Cu_In2.Cu": 0.225,
    "F.Cu_B.Cu": 1.2
}

def mm_to_ps(layer: str, mm: float) -> float:
    return mm * (layer_delay[layer] / 10)

def get_track_flytime(tracks, netname):
    """returns the flytime of all tracks in the net
    """
    flytime = 0
    for track in tracks:
        if track.GetNet().GetNetname() == netname and type(track) == pcbnew.PCB_TRACK:
            layer = track.GetLayerName()
            length = pcbnew.ToMM(track.GetLength())
            flytime += mm_to_ps(layer, length)
    return flytime

def get_net_layers(tracks, netname):
    """returns a list of layers used by the net
    """
    layers = []
    for track in tracks:
        if track.GetNet().GetNetname() == netname and type(track) == pcbnew.PCB_TRACK:
            layer = track.GetLayerName()
            if layer not in layers:
                layers.append(layer)
    return layers

def get_via_count(tracks, netname):
    """returns the number of vias in the net
    """
    via_count = 0
    for track in tracks:
        if track.GetNet().GetNetname() == netname and type(track) == pcbnew.PCB_VIA:
            via_count += 1
    return via_count

def get_via_delay(tracks, netname):
    """returns the via delay of the net
    """
    via_delay = 0
    vias = get_via_count(tracks, netname)
    layers = get_net_layers(tracks, netname)
    for i in range(len(layers)-1):
        via_delay += via_delays[layers[i] + "_" + layers[i+1]]
    return via_delay*2

def get_track_length(tracks, netname):
    """returns the length of all tracks in the net
    """
    length = 0
    for track in tracks:
        if track.GetNet().GetNetname() == netname and type(track) == pcbnew.PCB_TRACK:
            length += pcbnew.ToMM(track.GetLength())
    return length

def get_via_length(tracks, netname):
    """returns the length of all vias in the net
    """
    length = 0
    vias = get_via_count(tracks, netname)
    layers = get_net_layers(tracks, netname)
    for i in range(len(layers)-1):
        length += via_lengths[layers[i] + "_" + layers[i+1]]
    return length*2

def update_row(row, tracks):
    netname = NET_PREFIX + row[1].value

    vias = get_via_count(tracks, netname)
    row[2].value = vias

    track_flytime = get_track_flytime(tracks, netname)
    row[3].value = round(track_flytime, 1)

    via_delay = get_via_delay(tracks, netname)
    row[4].value = round(via_delay, 1)

    package_delay = row[5].value
    extra_delay = row[6].value

    total_delay = track_flytime + package_delay + extra_delay
    row[7].value = round(total_delay, 1)

    layers = get_net_layers(tracks, netname)
    row[8].value = ", ".join(layers)

    track_length = get_track_length(tracks, netname)
    row[9].value = round(track_length, 2)

    via_length = get_via_length(tracks, netname)
    row[10].value = round(via_length, 2)

    total_length = track_length + via_length
    row[11].value = round(total_length, 2)

def update_flytimes(ws, pcb, tracks):
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        update_row(row, tracks)

if __name__ == "__main__":
    wb = load_workbook("DDR_flytimes.xlsx")
    ws = wb.active
    pcb = pcbnew.LoadBoard("../qsbc.kicad_pcb")
    pcb.BuildListOfNets()
    tracks = pcb.GetTracks()
    update_flytimes(ws, pcb, tracks)
    wb.save("DDR_flytimes.xlsx")
