#!/usr/bin/env python3

import sys
import json
from turtle import width

import pcbnew

from openpyxl import Workbook, load_workbook

class TrackDelayNotFoundException(Exception):
    def __init__(self, layer: str, width: float):
        message = f"Track delay not found for track on layer {layer} with width {width}mm"
        super().__init__(message)

class ViaDelayNotFoundException(Exception):
    def __init__(self, width: float, drill: float, start: str, end: str) -> None:
        message = f"Via delay not found for via with {width} width, {drill} drill, {start} start, {end} end"
        super().__init__(message)

class ViaLayersException(Exception):
    def __init__(self, via: pcbnew.PCB_VIA):
        message = f"Via on net {via.GetNetname()} has unknown start and stop layers"
        super().__init__(message)

class UnhandeledElementException(Exception):
    def __init__(self, element: pcbnew.BOARD_CONNECTED_ITEM):
        super().__init__(f"Unhandeled element: {element}")

class NetNotFoundException(Exception):
    def __init__(self, netname: str):
        super().__init__(f"Net {netname} not found")

class FlyData:
    def __init__(self, filename: str):
        self.flytime_data = json.loads(open(filename).read())

    def get_track_pspcm(self, layer: str, width: float) -> float:
        for entry in self.flytime_data["tracks"][layer]:
            if entry["width"] == width:
                return entry["ps_per_cm"]
        raise TrackDelayNotFoundException(layer, width)

    def get_via_delay(self, width: float, drill: float, start: str, end: str):
        for entry in self.flytime_data["vias"]:
            if entry["width"] == width and entry["drill"] == drill and entry["start"] == start and entry["end"] == end:
                return entry["delay"]
        raise ViaDelayNotFoundException(width, drill, start, end)

    def get_via_height(self, width: float, drill: float, start: str, end: str):
        for entry in self.flytime_data["vias"]:
            if entry["width"] == width and entry["drill"] == drill and entry["start"] == start and entry["end"] == end:
                return entry["height"]
        raise ViaDelayNotFoundException(width, drill, start, end)

class FlyTools:

    def __init__(self, board: pcbnew.BOARD):
        self.board = board
        self.flydata = FlyData("flytime_info.json")
        self.board.BuildListOfNets()
        self.nettracks = self.board.GetTracks()
        self.tracks = []
        self.vias = []
        for track in self.nettracks:
            if type(track) == pcbnew.PCB_TRACK:
                self.tracks.append(track)
            elif type(track) == pcbnew.PCB_VIA:
                self.vias.append(track)
        self.nets = self.board.GetNetsByNetcode()

    @staticmethod
    def mm_to_ps(pspcm: float, mm: float) -> float:
        return mm * (pspcm / 10)

    def via_start_end(self, via: pcbnew.PCB_VIA) -> tuple:
        """Gets the first and last connected layer of a via"""
        toplayer = None
        bottomlayer = None
        for layer in range(via.TopLayer(), via.BottomLayer() + 1):
            if (self.board.GetConnectivity().IsConnectedOnLayer(via, layer)):
                if toplayer is None:
                    toplayer = layer
                else:
                    bottomlayer = layer
        if toplayer is None or bottomlayer is None:
            raise ViaLayersException(via)
        return (toplayer, bottomlayer)

    def get_element_delay(self, element: pcbnew.BOARD_CONNECTED_ITEM) -> float:
        """Get the delay of a single element in ps"""
        if type(element) == pcbnew.PCB_TRACK:
            return self.get_track_delay(element)
        elif type(element) == pcbnew.PCB_VIA:
            return self.get_via_delay(element)
        else:
            raise UnhandeledElementException(element)

    def get_track_delay(self, track: pcbnew.PCB_TRACK) -> float:
        layer = track.GetLayerName()
        length = pcbnew.ToMM(track.GetLength())
        width = pcbnew.ToMM(track.GetWidth())
        pspcm = self.flydata.get_track_pspcm(layer, width)
        return self.mm_to_ps(pspcm, length)

    def get_track_lengths(self, net: pcbnew.NETINFO_ITEM) -> float:
        """Get the length of all tracks on a net in mm"""
        length = 0
        for track in self.tracks:
            if track.GetNetCode() == net.GetNetCode():
                length += pcbnew.ToMM(track.GetLength())
        return length

    def get_via_delay(self, via: pcbnew.PCB_VIA) -> float:
        start, end = self.via_start_end(via)
        start = pcbnew.LayerName(start)
        end = pcbnew.LayerName(end)
        drill = pcbnew.ToMM(via.GetDrillValue())
        width = pcbnew.ToMM(via.GetWidth())
        delay = self.flydata.get_via_delay(width, drill, start, end)
        return delay

    def get_via_length(self, via: pcbnew.PCB_VIA) -> float:
        """Get the length of a via in mm"""
        #stackup = self.board.GetDesignSettings().GetStackupDescriptor()
        width = pcbnew.ToMM(via.GetWidth())
        drill = pcbnew.ToMM(via.GetDrillValue())
        start, end = self.via_start_end(via)
        start = pcbnew.LayerName(start)
        end = pcbnew.LayerName(end)
        #return stackup.getLayerDistance(start, end)
        return self.flydata.get_via_height(width, drill, start, end)


    def get_via_lengths(self, net: pcbnew.NETINFO_ITEM) -> float:
        """Get the length of all vias on a net in mm"""
        length = 0
        for via in self.vias:
            if via.GetNetCode() == net.GetNetCode():
                length += self.get_via_length(via)
        return length

    def get_items(self, net: pcbnew.NETINFO_ITEM) -> list[pcbnew.BOARD_CONNECTED_ITEM]:
        """Get all items on a net by shortname"""
        items = []
        for track in self.nettracks:
            if track.GetNetCode() == net.GetNetCode():
                items.append(track)
        return items

    def get_net_tracks_delay(self, net: pcbnew.NETINFO_ITEM) -> float:
        """Get the delay of all tracks on a net in ps"""
        delay = 0
        for track in self.tracks:
            if track.GetNetCode() == net.GetNetCode() and type(track) == pcbnew.PCB_TRACK:
                delay += self.get_element_delay(track)
        return delay

    def get_net_via_delays(self, net: pcbnew.NETINFO_ITEM) -> float:
        """Get the delay of all vias on a net in ps"""
        delay = 0
        for element in self.vias:
            if element.GetNetCode() == net.GetNetCode():
                delay += self.get_element_delay(element)
        return delay

    def get_num_vias(self, net: pcbnew.NETINFO_ITEM) -> int:
        """Get the number of vias on a net"""
        num = 0
        for element in self.vias:
            if element.GetNetCode() == net.GetNetCode():
                num += 1
        return num

    def get_layers(self, net: pcbnew.NETINFO_ITEM) -> list[str]:
        """Get the layers of a net"""
        layers = []
        for track in self.tracks:
            if track.GetNetCode() == net.GetNetCode():
                name = track.GetLayerName()
                if name not in layers:
                    layers.append(name)
        return layers

    def get_net_delay(self, net: pcbnew.NETINFO_ITEM) -> float:
        """Get the delay of a net in ps"""
        delay = 0
        for element in self.get_items(net):
            delay += self.get_element_delay(element)
        return delay

    def name_to_net(self, name: str) -> pcbnew.NETINFO_ITEM:
        """Get a net by name"""
        pcbnew.NETINFO_ITEM(self.board, name)
        raise NetNotFoundException(name)

    def shortname_to_net(self, shortname: str) -> pcbnew.NETINFO_ITEM:
        """Get a net by shortname"""
        for net in self.nets.items():
            if net[1].GetShortNetname() == shortname:
                return net[1]
        raise NetNotFoundException(shortname)

class FlySheet:

    def __init__(self, filename: str, flytools: FlyTools):
        self.filename = filename
        self.flytools = flytools
        self.workbook = load_workbook(filename)
        self.sheet = self.workbook.active

    def write(self, row: int, column: int, value):
        self.sheet.cell(row=row, column=column, value=value)

    def save(self):
        self.workbook.save(self.filename)

    def getColByName(self, name: str) -> int:
        for col in range(1, self.sheet.max_column + 1):
            if self.sheet.cell(row=1, column=col).value == name:
                return col
        raise Exception(f"Column {name} not found")

    def getFloatValue(self, row: int, column: str) -> float:
        col = self.getColByName(column)
        val = self.sheet.cell(row=row, column=col).value
        if val is None:
            return 0.0
        return float(val)

    def getValue(self, row: int, column: str) -> str:
        col = self.getColByName(column)
        val = self.sheet.cell(row=row, column=col).value
        return val

    def setValue(self, row: int, col: str, value: str or float):
        col = self.getColByName(col)
        self.write(row, col, value)

    def updateRow(self, row: int):

        netname = self.getValue(row, "Net")
        if netname is None:
            return
        net = self.flytools.shortname_to_net(netname)

        r = 2

        num_vias = self.flytools.get_num_vias(net)
        layers = self.flytools.get_layers(net)
        track_len = round(self.flytools.get_track_lengths(net), r)
        via_len = round(self.flytools.get_via_lengths(net), r)
        total_len = track_len + via_len
        track_delay = round(self.flytools.get_net_tracks_delay(net), r)
        via_delay = round(self.flytools.get_net_via_delays(net), r)

        package_delay = self.getFloatValue(row, "Package Delay")
        extra_delay = self.getFloatValue(row, "Extra Delay")

        total_delay = track_delay + via_delay + package_delay + extra_delay

        self.setValue(row, "Vias", num_vias)
        self.setValue(row, "Layers", ", ".join(layers))
        self.setValue(row, "Track Length", track_len)
        self.setValue(row, "Via Length", via_len)
        self.setValue(row, "Total Length", total_len)
        self.setValue(row, "Track Delay", track_delay)
        self.setValue(row, "Via Delay", via_delay)
        self.setValue(row, "Total Delay", total_delay)

    def update(self):
        for row in self.sheet.iter_rows(min_row=2):
            self.updateRow(row[0].row)

    def updateAll(self):
        for ws in self.workbook.worksheets:
            self.sheet = ws
            self.update()

if __name__ == "__main__":

    if len(sys.argv) < 2:
        print("Usage: gen_flytimes.py <kicad_pcb> <spreadsheet>")
        sys.exit(1)

    board = pcbnew.LoadBoard(sys.argv[1])
    flytools = FlyTools(board)
    flysheet = FlySheet(sys.argv[2], flytools)
    flysheet.updateAll()
    flysheet.save()
